#!/usr/bin/env python3

# Import modules
import sys
import os
import json
import pyzabbix
import getpass
import openpyxl
import platform
import subprocess
import time
import datetime
import collections

from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import (PatternFill, Border, Side, Alignment, Font, GradientFill, colors, Color)
from progress.bar import IncrementalBar


# Input for 1 Zabbix API, Username and Password, timeFrom, timeTill
apiPathNew = input(f"Set 1 Zabbix API address please:\n")
apiUsernameNew = input("Type 1 Zabbix API username please: ")
apiPasswordNew = getpass.getpass("Type 1 Zabbix API user password please: ")

## Input for 2 Zabbix API, Username and Password, timeFrom, timeTill
apiPathOld = input(f"Set 2 Zabbix API address please:\n")
apiUsernameOld = input("Type 2 Zabbix API username please: ")
apiPasswordOld = getpass.getpass("Type 2 Zabbix API user password please: ")

# Auth complete (New)
try:
    zabbixApi = pyzabbix.ZabbixAPI(apiPathNew)
except Exception:
    print("Incorrect API path... Try again")	
try:
    zabbixApi.login(user=apiUsernameNew, password=apiPasswordNew)
except Exception:
    print("Incorrect login... Try again")
try:
    zabbixApi.auth
except Exception:
    print("Failed to authorize... Try again")	

# Get API name - zabbix server address
zabbixApiName = apiPathNew.lstrip('https://')
zabbixApiName = zabbixApiName.rstrip('/zabbix/api_jsonrpc.php')


# Frameworks for ICMP ping history getting 
## If you will type nothing - these will be marked as default last 30 days history
fromTime = input(f"|- Set date/time from (ex. 20/11/2022 10:00), default: 30 days ago -|\n")
tillTime = input(f"|- Set date/time till (ex. 26/11/2022 12:00), default: now -|\n")

if tillTime == "":
    tillTime = int(time.mktime(datetime.now().timetuple()))
else:
    struct_time = time.strptime(tillTime, '%d/%m/%Y %H:%M')
    tillTime = int(time.mktime(struct_time))

if fromTime == "":
    fromTime = int(tillTime - 60 * 60 * 24)
else:
    struct_time = time.strptime(fromTime, '%d/%m/%Y %H:%M')
    fromTime = int(time.mktime(struct_time))


# Start time
if platform.system().lower()!='windows':
    start_time = time.time() 
else:
    start_time = time.clock()


### Part I - Zabbix 1 ###

if apiPathNew != None:

# Excel files folder creation
    currentFolder = os.getcwd()
    resultFolder = 'Excel'

# Change working directory to "Excel" if it exists, else - make directory and change current directory
    try: 
        os.chdir("./Excel/")
    except Exception:
        os.mkdir(resultFolder, 0o700)
        os.chdir("./Excel/")

# Preparing JSON data from Zabbix API
    itemFilter = {'key_': 'icmpping'}
    itemFilterOS = {'key_': 'system.sw.os'}
   
    rawHostInfo = zabbixApi.host.get(filter={'host': "Zabbix server"}, output = ['hostid','host','name','status'], \
    selectInterfaces=['ip','port','dns', 'type'], selectGroups=['groupid', 'name'], \
    selectParentTemplates=['templateid', 'name'], \
    selectInventory=['alias','chassis','hardware_full','host_netmask','host_networks','macaddress_a','model','name','os_full','serialno_a','type_full'], \
    selectItems=['itemid','delay','interfaceid','key_','name','type','value_type','description','history','trends','state','status'], \
    selectTriggers=['triggerid','description','expression','priority','state','status','recovery_mode','recovery_expression','manual_close'])
    
    JSrequest = json.dumps(rawHostInfo)
    
# Serializing json
    json_object = json.dumps(rawHostInfo, indent=4)
 
# Writing to new.json
    with open("new.json", "w") as outfile:
        outfile.write(json_object)
    
    data = json.loads(JSrequest)

    print("|----- Get JSON data - done -----|")

# Create Workbook and active worksheet named according to the Zabbix API Server
    wb = Workbook()
    ws = wb.active
    firstSheetName = wb.worksheets[0]

    if ws.title != zabbixApiName:
        wb.create_sheet(f"{zabbixApiName}")
        ws = wb[f"{zabbixApiName}"]
        ws.title = zabbixApiName
    else:
        ws.title = zabbixApiName

# Deleting of the default worksheet
    wb.remove(firstSheetName)

# Creation of the table title
    ws['A1'].value = f"Hosts report from Zabbix Server: {zabbixApiName}"

    ws.append(["Hostid"] + ["Host"] + ["Name"] + ["Status"] + ["IP"] + ["Port"] + ["Type"] + ["OS"] + ["Avail"] + ["SLA (ICMP ping)"] + ["Inventory"] + ["Group"] + ["Template"] + ["Item"] + ["Trigger"])

# Prepare and combine data
    x = 0
    co = 0
    osValAll = str()
    hostCount = len(data)
    bar = IncrementalBar('Collecting info from new Zabbix server', max = len(data))
    groups = str()
    templates = str()
    inventory = str()
    itemsStr = str()
    triggersStr = str()

# General information
    for x in data:
        hostid = int(x['hostid'])
        host = str(x['host'])
        name = str(x['name']) 
        status = int(x['status']) 
        ip = str(x['interfaces'][0]['ip']) 
        port = int(x['interfaces'][0]['port']) 
        type = int(x['interfaces'][0]['type']) 

# Host groups (limited by 10 host groups, you can add more if necessary)
        groupId1 = str(x['groups'][0]['groupid'])
        groupName1 = str(x['groups'][0]['name'])
        try:
            groupId2 = str(x['groups'][1]['groupid'])
            groupName2 = str(x['groups'][1]['name'])
        except Exception:
            groupId2 = str("")
            groupName2 = str("")  
        try:
            groupId3 = str(x['groups'][2]['groupid'])
            groupName3 = str(x['groups'][2]['name'])
        except Exception:
            groupId3 = str("")
            groupName3 = str("") 
        try:
            groupId4 = str(x['groups'][3]['groupid'])
            groupName4 = str(x['groups'][3]['name'])
        except Exception:
            groupId4 = str("")
            groupName4 = str("")   
        try:
            groupId5 = str(x['groups'][4]['groupid'])
            groupName5 = str(x['groups'][4]['name'])
        except Exception:
            groupId5 = str("")
            groupName5 = str("")  
        try:
            groupId6 = str(x['groups'][5]['groupid'])
            groupName6 = str(x['groups'][5]['name'])
        except Exception:
            groupId6 = str("")
            groupName6 = str("") 
        try:
            groupId7 = str(x['groups'][6]['groupid'])
            groupName7 = str(x['groups'][6]['name'])
        except Exception:
            groupId7 = str("")
            groupName7 = str("") 
        try:
            groupId8 = str(x['groups'][7]['groupid'])
            groupName8 = str(x['groups'][7]['name'])
        except Exception:
            groupId8 = str("")
            groupName8 = str("")  
        try:
            groupId9 = str(x['groups'][8]['groupid'])
            groupName9 = str(x['groups'][8]['name'])
        except Exception:
            groupId9 = str("")
            groupName9 = str("") 
        try:
            groupId10 = str(x['groups'][9]['groupid'])
            groupName10 = str(x['groups'][9]['name'])
        except Exception:
            groupId10 = str("")
            groupName10 = str("") 
        try:
            groupId11 = str(x['groups'][10]['groupid'])
            groupName11 = str(x['groups'][10]['name'])
        except Exception:
            groupId11 = str("")
            groupName11 = str("")

        groups = f"{groupId1} {groupName1}\n{groupId2} {groupName2}\n{groupId3} {groupName3}\n{groupId4} {groupName4}\n\
        {groupId5} {groupName5}\n{groupId6} {groupName6}\n{groupId7} {groupName7}\n{groupId8} {groupName8}\n{groupId9} {groupName9}\n\
        {groupId10} {groupName10}\n{groupId11} {groupName11}"

# Get hosts templates info, 15 templates (you can add more)
        try:
            templateId1 = str(x['parentTemplates'][0]['templateid'])
            templateName1 = str(x['parentTemplates'][0]['name'])
        except Exception:
            templateId1 = str("")
            templateName1 = str("")
        try:
            templateId2 = str(x['parentTemplates'][1]['templateid'])
            templateName2 = str(x['parentTemplates'][1]['name'])
        except Exception:
            templateId2 = str("")
            templateName2 = str("")
        try:
            templateId3 = str(x['parentTemplates'][2]['templateid'])
            templateName3 = str(x['parentTemplates'][2]['name'])
        except Exception:
            templateId3 = str("")
            templateName3 = str("")
        try:
            templateId4 = str(x['parentTemplates'][3]['templateid'])
            templateName4 = str(x['parentTemplates'][3]['name'])
        except Exception:
            templateId4 = str("")
            templateName4 = str("")
        try:
            templateId5 = str(x['parentTemplates'][4]['templateid'])
            templateName5 = str(x['parentTemplates'][4]['name'])
        except Exception:
            templateId5 = str("")
            templateName5 = str("")
        try:
            templateId6 = str(x['parentTemplates'][5]['templateid'])
            templateName6 = str(x['parentTemplates'][5]['name'])
        except Exception:
            templateId6 = str("")
            templateName6 = str("")
        try:
            templateId7 = str(x['parentTemplates'][6]['templateid'])
            templateName7 = str(x['parentTemplates'][6]['name'])
        except Exception:
            templateId7 = str("")
            templateName7 = str("")
        try:
            templateId8 = str(x['parentTemplates'][7]['templateid'])
            templateName8 = str(x['parentTemplates'][7]['name'])
        except Exception:
            templateId8 = str("")
            templateName8 = str("")
        try:
            templateId9 = str(x['parentTemplates'][8]['templateid'])
            templateName9 = str(x['parentTemplates'][8]['name'])
        except Exception:
            templateId9 = str("")
            templateName9 = str("")
        try:
            templateId10 = str(x['parentTemplates'][9]['templateid'])
            templateName10 = str(x['parentTemplates'][9]['name'])
        except Exception:
            templateId10 = str("")
            templateName10 = str("")
        try:
            templateId11 = str(x['parentTemplates'][10]['templateid'])
            templateName11 = str(x['parentTemplates'][10]['name'])
        except Exception:
            templateId11 = str("")
            templateName11 = str("")
        try:
            templateId12 = str(x['parentTemplates'][11]['templateid'])
            templateName12 = str(x['parentTemplates'][11]['name'])
        except Exception:
            templateId12 = str("")
            templateName12 = str("")
        try:
            templateId13 = str(x['parentTemplates'][12]['templateid'])
            templateName13 = str(x['parentTemplates'][12]['name'])
        except Exception:
            templateId13 = str("")
            templateName13 = str("")
        try:
            templateId14 = str(x['parentTemplates'][13]['templateid'])
            templateName14 = str(x['parentTemplates'][13]['name'])
        except Exception:
            templateId14 = str("")
            templateName14 = str("")
        try:
            templateId15 = str(x['parentTemplates'][14]['templateid'])
            templateName15 = str(x['parentTemplates'][14]['name'])
        except Exception:
            templateId15 = str("")
            templateName15 = str("")
            
        templates = f"{templateId1} {templateName1}\n{templateId2} {templateName2}\n{templateId3} {templateName3}\n{templateId4} {templateName4}\n\
        {templateId5} {templateName5}\n{templateId6} {templateName6}\n{templateId7} {templateName7}\n{templateId8} {templateName8}\n\
        {templateId9} {templateName9}\n{templateId10} {templateName10}\n{templateId11} {templateName11}\n{templateId12} {templateName12}\n\
        {templateId13} {templateName13}\n{templateId14} {templateName14}\n{templateId15} {templateName15}"      

# Inventory        
        try:
            alias = str(x['inventory']['alias'])
        except Exception:
            alias = str("")
        try:
            chassis = str(x['inventory']['chassis'])
        except Exception:
            chassis = str("")
        try:
            hardware = str(x['inventory']['hardware_full'])
        except Exception:
            hardware = str("")
        try:
            netmask = str(x['inventory']['host_netmask'])
        except Exception:
            netmask = str("")          
        try:
            networks = str(x['inventory']['host_networks'])
        except Exception:
            networks = str("")
        try:
            macaddress = str(x['inventory']['macaddress_a'])
        except Exception:
            macaddress = str("")
        try:
            model = str(x['inventory']['model'])
        except Exception:
            model = str("")
        try:
            name = str(x['inventory']['name'])
        except Exception:
            name = str("")            
        try:
            os_full = str(x['inventory']['os_full'])
        except Exception:
            os_full = str("")
        try:
            serialno = str(x['inventory']['serialno_a'])
        except Exception:
            serialno = str("")
        try:
            type_full = str(x['inventory']['type_full'])
        except Exception:
            type_full = str("")
        inventory = f"{name}\n{model}\n{alias}\n{chassis}\n{hardware}\n{netmask}\n{networks}\n{macaddress}\n{os_full}\n{serialno}\n{type_full}"

# Items   
        itemsAll = zabbixApi.item.get(host=x['host'], output='extend', selectHosts=['host','name'], \
            selectTriggers=['triggerid','description','expression','priority','state','status','recovery_mode','recovery_expression','manual_close'])

        itemsCount = len(itemsAll)

        for c in range(itemsCount):
            try:
                itemsId = str(itemsAll[c]['itemid'])
            except Exception:
                itemsId = str("")
            try:
                itemsname = str(itemsAll[c]['name'])
            except Exception:
                itemsname = str("")
            try:
                itemskey = str(itemsAll[c]['key_'])
            except Exception:
                itemskey = str("")
            try:
                itemsinterfaceid = str(itemsAll[c]['interfaceid'])
            except Exception:
                itemsinterfaceid = str("")
            try:
                itemstype = str(itemsAll[c]['type'])
            except Exception:
                itemstype = str("")
            try:
                itemsvalue_type = str(itemsAll[c]['value_type'])
            except Exception:
                itemsvalue_type = str("")
            try:
                itemsdescription = str(itemsAll[c]['description'])
            except Exception:
                itemsdescription = str("")
            try:
                itemshistory = str(itemsAll[c]['history'])
            except Exception:
                itemshistory = str("")
            try:
                itemstrends = str(itemsAll[c]['trends'])
            except Exception:
                itemstrends = str("")
            try:
                itemsstate = str(itemsAll[c]['state'])
            except Exception:
                itemsstate = str("")
            try:
                itemsstatus = str(itemsAll[c]['status'])
            except Exception:
                itemsstatus = str("")

            itemsStr += f"Item Id: {itemsId} name: {itemsname} key: {itemskey}\nInterface Id: {itemsinterfaceid} Type: {itemstype} Value Type: {itemsvalue_type}\n\
Desc: {itemsdescription}\nHistory: {itemshistory} Trends: {itemstrends} State: {itemsstate} Status: {itemsstatus}\n\n"

# Triggers
            triggerCount = int(len(itemsAll[c]['triggers']))
            if triggerCount != 0:
                triggerCount = triggerCount
            else:
                triggerCount = 1
            for tr in range(len(itemsAll[c]['triggers'])):
                try:
                    triggerId = itemsAll[c]['triggers'][tr]['triggerid']
                except Exception:
                    triggerId = str("")
                try:
                    triggername = itemsAll[c]['triggers'][tr]['name']
                except Exception:
                    triggername = str("")
                try:
                    triggerkey = itemsAll[c]['triggers'][tr]['key_']
                except Exception:
                    triggerkey = str("")
                try:
                    triggerdelay = itemsAll[c]['triggers'][tr]['delay']
                except Exception:
                    triggerdelay = str("")
                try:
                    triggerinterfaceid = itemsAll[c]['triggers'][tr]['interfaceid']
                except Exception:
                    triggerinterfaceid = str("")
                try:
                    triggertype = itemsAll[c]['triggers'][tr]['type']
                except Exception:
                    triggertype = str("")
                try:
                    triggervalue_type = itemsAll[c]['triggers'][tr]['value_type']
                except Exception:
                    triggervalue_type = str("")
                try:
                    triggerdescription = itemsAll[c]['triggers'][tr]['description']
                except Exception:
                    triggerdescription = str("")
                try:
                    triggerhistory = itemsAll[c]['triggers'][tr]['history']
                except Exception:
                    triggerhistory = str("")
                try:
                    triggertrends = itemsAll[c]['triggers'][tr]['trends']
                except Exception:
                    triggertrends = str("")
                try:
                    triggerstate = itemsAll[c]['triggers'][tr]['state']
                except Exception:
                    triggerstate = str("")
                try:
                    triggerstatus = itemsAll[c]['triggers'][tr]['status']
                except Exception:
                    triggerstatus = str("")

                triggersStr += f"Trigger Id: {triggerId} Name: {triggername} Key: {triggerkey}\n\
Delay: {triggerdelay} Interface Id: {triggerinterfaceid} Type: {triggertype}\nDesc: {triggerdescription}\n\
History: {triggerhistory} Trends: {triggertrends} State: {triggerstate} Status: {triggerstatus}\n\n"
       
# Ping
        param = '-n' if platform.system().lower()=='windows' else '-c'
        command = ['ping', param, '1', ip]
        try:
            result = subprocess.run(command, check=True, shell=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            avail = "available"
        except Exception:
            avail = "unreachable"

# SLA
        items = zabbixApi.item.get(filter=itemFilter, host=x['host'], output='extend', selectHosts=['host','name'])
        baseList = list()
        complList = list()
        numList = list()
        for item in items:
            values = zabbixApi.history.get(itemids=item['itemid'], time_from=fromTime, time_till = tillTime, history=item['value_type'])
            for historyValue in values:
                ed = list(historyValue['value'])
                baseList.append(ed) 
            for l in baseList:
                complList +=  l
        for i in complList:
            number = int(i)
            if number is not None:
                numList.append(number)
            else:
                numList.append(0)
        try:
            hostItemValueLength = len(numList)
        except Exception:
            hostItemValueLength = 0
        try:
            hostItemValueSum = sum(numList)
        except Exception:
            hostItemValueSum = 0
        try:
            sla = (hostItemValueSum/hostItemValueLength) * 100
        except Exception:
            sla = "uncomputed"

# OS check
        itemOS = zabbixApi.item.get(filter=itemFilterOS, host=x['host'], output='extend', selectHosts=['host','name'])
        try:
            for itemOsValue in itemOS:
                osVal = str(itemOsValue['lastvalue'])
                osVal = osVal.split(" ")
                osVal1 = str(osVal[0])
                osVal2 = str(osVal[1])
                osVal3 = str(osVal[2])
                osValAll = str(osVal1 + " " + osVal2 + " " + osVal3)
        except Exception:
            osValAll = "Undefined"

# Progress bar
        bar.next()

# Add all info into report
        ws.append([hostid] + [host] + [name] + [status] + [ip] + [port] + [type] + [osValAll] + [avail] + [sla] + [inventory] + [groups] + [templates] + [itemsStr] + [triggersStr])

    bar.finish()

    print("|----- Collect info and add into report - done -----|")

# Let's configurate the report
# Merge cells for title
    maxCol = ws.max_column
    minRow = ws.min_row
    maxRow = ws.max_row
    maxColLetter = get_column_letter(maxCol)
    refCell = maxColLetter+str(maxRow)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=maxCol)

### Styles ###
# Column width
    for col in ws.iter_cols(min_row=2, max_row=hostCount+2, min_col=1, max_col=10):
         max_length = 10
         column = col[0].column_letter
         for cell in col:
             try: 
                 if len(str(cell.value)) > max_length:
                     #max_length = len(str(cell.value))
                     max_length = 15
             except:
                 continue
         adjusted_width = (max_length)
         ws.column_dimensions[column].width = adjusted_width

    for col in ws.iter_cols(min_row=2, max_row=hostCount+2, min_col=11, max_col=13):
         max_length = 15
         column = col[0].column_letter
         for cell in col:
             try: 
                 if len(str(cell.value)) > max_length:
                     #max_length = len(str(cell.value))
                     max_length = 30
             except:
                 continue
         adjusted_width = (max_length)
         ws.column_dimensions[column].width = adjusted_width


    for col in ws.iter_cols(min_row=2, max_row=hostCount+2, min_col=14, max_col=15):
         max_length = 50
         column = col[0].column_letter
         for cell in col:
             try: 
                 if len(str(cell.value)) > max_length:
                     #max_length = len(str(cell.value))
                     max_length = 100
             except:
                 continue
         adjusted_width = (max_length)
         ws.column_dimensions[column].width = adjusted_width


# Alignment style
    alignment=Alignment(
                    horizontal='left',
                    vertical='top',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=False,
                    indent=0
                       )

# Font styles
    fontTitle = Font(
            name='Bahnschrift SemiBold',
            size=14,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000'
                )

    fontHeaders = Font(
            name='Bahnschrift',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000'
                )

    fontCells = Font(
            name='Bahnschrift Light',
            size=9,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000'
                )

# Fill cells style
    fillTitle = PatternFill(fill_type='solid', fgColor='FFCBA4')
    fillHeaders = PatternFill(fill_type='solid', fgColor='96c8a2')
    fillCells = PatternFill(fill_type='solid', fgColor='addfad')
    fillAvail = PatternFill(fill_type='solid', fgColor='32CD32')
    fillUnreach = PatternFill(fill_type='solid', fgColor='CD5C5C')

# Borders style
    borderTitle = Side(border_style="thick", color="2f4f4f")
    borderHeader = Side(border_style="medium", color="2f4f4f")
    borderCells = Side(border_style="thin", color="2f4f4f")

# Title cell
    ws['A1'].font = fontTitle
    ws['A1'].fill = fillTitle
    ws['A1'].alignment = alignment

# Headers
    for cells in ws.iter_cols(min_row=2, max_row=2):
        for cell in cells:
            cell.font = fontHeaders
            cell.fill = fillHeaders
            cell.alignment = alignment
            cell.border = Border(top=borderHeader, bottom=borderHeader, left=borderHeader, right=borderHeader) 

# Fill cells skipping one
    for rows in range(minRow + 2, maxRow + 2, 2):
        for cells in ws.iter_cols(min_row=rows, max_row=rows):
            for cell in cells:
                cell.fill = fillCells

# Data cells
    for cells in ws.iter_cols(min_row=3, max_row=hostCount+2):
        for cell in cells:
            styleObj = cell.coordinate
            ws[f'{styleObj}'].alignment = alignment
            ws[f'{styleObj}'].border = Border(top=borderCells, bottom=borderCells, left=borderCells, right=borderCells)
            if cell.value == "available":
                cell.fill = fillAvail
            elif cell.value == "unreachable":
                cell.fill = fillUnreach

    for i in range(1, ws.max_row+1):
        ws.row_dimensions[i].height = 30

    ws.page_setup.scale = 80
        
    print("|----- Configure report cells style - done -----|")

# Log out from Zabbix API
    zabbixApi.user.logout()

else:
    print("Incorrect input. Try again.")

#### Part II - Zabbix 2 ####

# Auth complete (Old)
try:
    zabbixApi2 = pyzabbix.ZabbixAPI(apiPathOld)
except Exception:
    print("Incorrect API path... Try again")	
try:
    zabbixApi2.login(user=apiUsernameOld, password=apiPasswordOld)
except Exception:
    print("Incorrect login... Try again")
try:
    zabbixApi2.auth
except Exception:
    print("Failed to authorize... Try again")	

# Get API name - zabbix server address
zabbixApiName2 = apiPathOld.lstrip('https://')
zabbixApiName2 = zabbixApiName2.rstrip('/zabbix/api_jsonrpc.php')

if apiPathOld != None:

# Preparing JSON data from Zabbix API
    itemFilter = {'key_': 'icmpping'}
    itemFilterOS = {'key_': 'system.sw.os'}
   
    rawHostInfo = zabbixApi2.host.get(filter={'host': "Zabbix server"}, output = ['hostid','host','name','status'], \
    selectInterfaces=['ip','port','dns', 'type'], selectGroups=['groupid', 'name'], \
    selectParentTemplates=['templateid', 'name'], \
    selectInventory=['alias','chassis','hardware_full','host_netmask','host_networks','macaddress_a','model','name','os_full','serialno_a','type_full'], \
    selectItems=['itemid','delay','interfaceid','key_','name','type','value_type','description','history','trends','state','status'], \
    selectTriggers=['triggerid','description','expression','priority','state','status','recovery_mode','recovery_expression','manual_close'])
    
    JSrequest = json.dumps(rawHostInfo)
    
# Serializing json
    json_object2 = json.dumps(rawHostInfo, indent=4)
 
# Writing to new.json
    with open("old.json", "w") as outfile:
        outfile.write(json_object2)
    
    data = json.loads(JSrequest)

    print("|----- Get JSON data - done -----|")

# Create Worksheet and active worksheet named according to the Zabbix API Server

    wb.create_sheet(f"{zabbixApiName2}")
    ws2 = wb[f"{zabbixApiName2}"]
    ws2.title = zabbixApiName2

# Creation of the table title
    ws2['A1'].value = f"Hosts report from Zabbix Server: {zabbixApiName2}"

    ws2.append(["Hostid"] + ["Host"] + ["Name"] + ["Status"] + ["IP"] + ["Port"] + ["Type"] + ["OS"] + ["Avail"] + ["SLA (ICMP ping)"] + ["Inventory"] + ["Group"] + ["Template"] + ["Item"] + ["Trigger"])

# Prepare and combine data
    x = 0
    co = 0
    osValAll = str()
    hostCount = len(data)
    bar = IncrementalBar('Collecting info from new Zabbix server', max = len(data))
    groups = str()
    templates = str()
    inventory = str()
    itemsStr = str()
    triggersStr = str()

# General information
    for x in data:
        hostid = int(x['hostid'])
        host = str(x['host'])
        name = str(x['name']) 
        status = int(x['status']) 
        ip = str(x['interfaces'][0]['ip']) 
        port = int(x['interfaces'][0]['port']) 
        type = int(x['interfaces'][0]['type']) 

# Host groups (limited by 10 host groups, you can add more if necessary)
        groupId1 = str(x['groups'][0]['groupid'])
        groupName1 = str(x['groups'][0]['name'])
        try:
            groupId2 = str(x['groups'][1]['groupid'])
            groupName2 = str(x['groups'][1]['name'])
        except Exception:
            groupId2 = str("")
            groupName2 = str("")  
        try:
            groupId3 = str(x['groups'][2]['groupid'])
            groupName3 = str(x['groups'][2]['name'])
        except Exception:
            groupId3 = str("")
            groupName3 = str("") 
        try:
            groupId4 = str(x['groups'][3]['groupid'])
            groupName4 = str(x['groups'][3]['name'])
        except Exception:
            groupId4 = str("")
            groupName4 = str("")   
        try:
            groupId5 = str(x['groups'][4]['groupid'])
            groupName5 = str(x['groups'][4]['name'])
        except Exception:
            groupId5 = str("")
            groupName5 = str("")  
        try:
            groupId6 = str(x['groups'][5]['groupid'])
            groupName6 = str(x['groups'][5]['name'])
        except Exception:
            groupId6 = str("")
            groupName6 = str("") 
        try:
            groupId7 = str(x['groups'][6]['groupid'])
            groupName7 = str(x['groups'][6]['name'])
        except Exception:
            groupId7 = str("")
            groupName7 = str("") 
        try:
            groupId8 = str(x['groups'][7]['groupid'])
            groupName8 = str(x['groups'][7]['name'])
        except Exception:
            groupId8 = str("")
            groupName8 = str("")  
        try:
            groupId9 = str(x['groups'][8]['groupid'])
            groupName9 = str(x['groups'][8]['name'])
        except Exception:
            groupId9 = str("")
            groupName9 = str("") 
        try:
            groupId10 = str(x['groups'][9]['groupid'])
            groupName10 = str(x['groups'][9]['name'])
        except Exception:
            groupId10 = str("")
            groupName10 = str("") 
        try:
            groupId11 = str(x['groups'][10]['groupid'])
            groupName11 = str(x['groups'][10]['name'])
        except Exception:
            groupId11 = str("")
            groupName11 = str("")

        groups = f"{groupId1} {groupName1}\n{groupId2} {groupName2}\n{groupId3} {groupName3}\n{groupId4} {groupName4}\n\
        {groupId5} {groupName5}\n{groupId6} {groupName6}\n{groupId7} {groupName7}\n{groupId8} {groupName8}\n{groupId9} {groupName9}\n\
        {groupId10} {groupName10}\n{groupId11} {groupName11}"

# Get hosts templates info, 15 templates (you can add more)
        try:
            templateId1 = str(x['parentTemplates'][0]['templateid'])
            templateName1 = str(x['parentTemplates'][0]['name'])
        except Exception:
            templateId1 = str("")
            templateName1 = str("")
        try:
            templateId2 = str(x['parentTemplates'][1]['templateid'])
            templateName2 = str(x['parentTemplates'][1]['name'])
        except Exception:
            templateId2 = str("")
            templateName2 = str("")
        try:
            templateId3 = str(x['parentTemplates'][2]['templateid'])
            templateName3 = str(x['parentTemplates'][2]['name'])
        except Exception:
            templateId3 = str("")
            templateName3 = str("")
        try:
            templateId4 = str(x['parentTemplates'][3]['templateid'])
            templateName4 = str(x['parentTemplates'][3]['name'])
        except Exception:
            templateId4 = str("")
            templateName4 = str("")
        try:
            templateId5 = str(x['parentTemplates'][4]['templateid'])
            templateName5 = str(x['parentTemplates'][4]['name'])
        except Exception:
            templateId5 = str("")
            templateName5 = str("")
        try:
            templateId6 = str(x['parentTemplates'][5]['templateid'])
            templateName6 = str(x['parentTemplates'][5]['name'])
        except Exception:
            templateId6 = str("")
            templateName6 = str("")
        try:
            templateId7 = str(x['parentTemplates'][6]['templateid'])
            templateName7 = str(x['parentTemplates'][6]['name'])
        except Exception:
            templateId7 = str("")
            templateName7 = str("")
        try:
            templateId8 = str(x['parentTemplates'][7]['templateid'])
            templateName8 = str(x['parentTemplates'][7]['name'])
        except Exception:
            templateId8 = str("")
            templateName8 = str("")
        try:
            templateId9 = str(x['parentTemplates'][8]['templateid'])
            templateName9 = str(x['parentTemplates'][8]['name'])
        except Exception:
            templateId9 = str("")
            templateName9 = str("")
        try:
            templateId10 = str(x['parentTemplates'][9]['templateid'])
            templateName10 = str(x['parentTemplates'][9]['name'])
        except Exception:
            templateId10 = str("")
            templateName10 = str("")
        try:
            templateId11 = str(x['parentTemplates'][10]['templateid'])
            templateName11 = str(x['parentTemplates'][10]['name'])
        except Exception:
            templateId11 = str("")
            templateName11 = str("")
        try:
            templateId12 = str(x['parentTemplates'][11]['templateid'])
            templateName12 = str(x['parentTemplates'][11]['name'])
        except Exception:
            templateId12 = str("")
            templateName12 = str("")
        try:
            templateId13 = str(x['parentTemplates'][12]['templateid'])
            templateName13 = str(x['parentTemplates'][12]['name'])
        except Exception:
            templateId13 = str("")
            templateName13 = str("")
        try:
            templateId14 = str(x['parentTemplates'][13]['templateid'])
            templateName14 = str(x['parentTemplates'][13]['name'])
        except Exception:
            templateId14 = str("")
            templateName14 = str("")
        try:
            templateId15 = str(x['parentTemplates'][14]['templateid'])
            templateName15 = str(x['parentTemplates'][14]['name'])
        except Exception:
            templateId15 = str("")
            templateName15 = str("")
            
        templates = f"{templateId1} {templateName1}\n{templateId2} {templateName2}\n{templateId3} {templateName3}\n{templateId4} {templateName4}\n\
        {templateId5} {templateName5}\n{templateId6} {templateName6}\n{templateId7} {templateName7}\n{templateId8} {templateName8}\n\
        {templateId9} {templateName9}\n{templateId10} {templateName10}\n{templateId11} {templateName11}\n{templateId12} {templateName12}\n\
        {templateId13} {templateName13}\n{templateId14} {templateName14}\n{templateId15} {templateName15}"      

# Inventory        
        try:
            alias = str(x['inventory']['alias'])
        except Exception:
            alias = str("")
        try:
            chassis = str(x['inventory']['chassis'])
        except Exception:
            chassis = str("")
        try:
            hardware = str(x['inventory']['hardware_full'])
        except Exception:
            hardware = str("")
        try:
            netmask = str(x['inventory']['host_netmask'])
        except Exception:
            netmask = str("")          
        try:
            networks = str(x['inventory']['host_networks'])
        except Exception:
            networks = str("")
        try:
            macaddress = str(x['inventory']['macaddress_a'])
        except Exception:
            macaddress = str("")
        try:
            model = str(x['inventory']['model'])
        except Exception:
            model = str("")
        try:
            name = str(x['inventory']['name'])
        except Exception:
            name = str("")            
        try:
            os_full = str(x['inventory']['os_full'])
        except Exception:
            os_full = str("")
        try:
            serialno = str(x['inventory']['serialno_a'])
        except Exception:
            serialno = str("")
        try:
            type_full = str(x['inventory']['type_full'])
        except Exception:
            type_full = str("")
        inventory = f"{name}\n{model}\n{alias}\n{chassis}\n{hardware}\n{netmask}\n{networks}\n{macaddress}\n{os_full}\n{serialno}\n{type_full}"

# Items   
        itemsAll = zabbixApi2.item.get(host=x['host'], output='extend', selectHosts=['host','name'], \
            selectTriggers=['triggerid','description','expression','priority','state','status','recovery_mode','recovery_expression','manual_close'])

        itemsCount = len(itemsAll)

        for c in range(itemsCount):
            try:
                itemsId = str(itemsAll[c]['itemid'])
            except Exception:
                itemsId = str("")
            try:
                itemsname = str(itemsAll[c]['name'])
            except Exception:
                itemsname = str("")
            try:
                itemskey = str(itemsAll[c]['key_'])
            except Exception:
                itemskey = str("")
            try:
                itemsinterfaceid = str(itemsAll[c]['interfaceid'])
            except Exception:
                itemsinterfaceid = str("")
            try:
                itemstype = str(itemsAll[c]['type'])
            except Exception:
                itemstype = str("")
            try:
                itemsvalue_type = str(itemsAll[c]['value_type'])
            except Exception:
                itemsvalue_type = str("")
            try:
                itemsdescription = str(itemsAll[c]['description'])
            except Exception:
                itemsdescription = str("")
            try:
                itemshistory = str(itemsAll[c]['history'])
            except Exception:
                itemshistory = str("")
            try:
                itemstrends = str(itemsAll[c]['trends'])
            except Exception:
                itemstrends = str("")
            try:
                itemsstate = str(itemsAll[c]['state'])
            except Exception:
                itemsstate = str("")
            try:
                itemsstatus = str(itemsAll[c]['status'])
            except Exception:
                itemsstatus = str("")

            itemsStr += f"Item Id: {itemsId} name: {itemsname} key: {itemskey}\nInterface Id: {itemsinterfaceid} Type: {itemstype} Value Type: {itemsvalue_type}\n\
Desc: {itemsdescription}\nHistory: {itemshistory} Trends: {itemstrends} State: {itemsstate} Status: {itemsstatus}\n\n"

# Triggers
            triggerCount = int(len(itemsAll[c]['triggers']))
            if triggerCount != 0:
                triggerCount = triggerCount
            else:
                triggerCount = 1
            for tr in range(len(itemsAll[c]['triggers'])):
                try:
                    triggerId = itemsAll[c]['triggers'][tr]['triggerid']
                except Exception:
                    triggerId = str("")
                try:
                    triggername = itemsAll[c]['triggers'][tr]['name']
                except Exception:
                    triggername = str("")
                try:
                    triggerkey = itemsAll[c]['triggers'][tr]['key_']
                except Exception:
                    triggerkey = str("")
                try:
                    triggerdelay = itemsAll[c]['triggers'][tr]['delay']
                except Exception:
                    triggerdelay = str("")
                try:
                    triggerinterfaceid = itemsAll[c]['triggers'][tr]['interfaceid']
                except Exception:
                    triggerinterfaceid = str("")
                try:
                    triggertype = itemsAll[c]['triggers'][tr]['type']
                except Exception:
                    triggertype = str("")
                try:
                    triggervalue_type = itemsAll[c]['triggers'][tr]['value_type']
                except Exception:
                    triggervalue_type = str("")
                try:
                    triggerdescription = itemsAll[c]['triggers'][tr]['description']
                except Exception:
                    triggerdescription = str("")
                try:
                    triggerhistory = itemsAll[c]['triggers'][tr]['history']
                except Exception:
                    triggerhistory = str("")
                try:
                    triggertrends = itemsAll[c]['triggers'][tr]['trends']
                except Exception:
                    triggertrends = str("")
                try:
                    triggerstate = itemsAll[c]['triggers'][tr]['state']
                except Exception:
                    triggerstate = str("")
                try:
                    triggerstatus = itemsAll[c]['triggers'][tr]['status']
                except Exception:
                    triggerstatus = str("")

                triggersStr += f"Trigger Id: {triggerId} Name: {triggername} Key: {triggerkey}\n\
Delay: {triggerdelay} Interface Id: {triggerinterfaceid} Type: {triggertype}\nDesc: {triggerdescription}\n\
History: {triggerhistory} Trends: {triggertrends} State: {triggerstate} Status: {triggerstatus}\n\n"
       
# Ping
        param = '-n' if platform.system().lower()=='windows' else '-c'
        command = ['ping', param, '1', ip]
        try:
            result = subprocess.run(command, check=True, shell=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            avail = "available"
        except Exception:
            avail = "unreachable"

# SLA
        items = zabbixApi2.item.get(filter=itemFilter, host=x['host'], output='extend', selectHosts=['host','name'])
        baseList = list()
        complList = list()
        numList = list()
        for item in items:
            values = zabbixApi2.history.get(itemids=item['itemid'], time_from=fromTime, time_till = tillTime, history=item['value_type'])
            for historyValue in values:
                ed = list(historyValue['value'])
                baseList.append(ed) 
            for l in baseList:
                complList +=  l
        for i in complList:
            number = int(i)
            if number is not None:
                numList.append(number)
            else:
                numList.append(0)
        try:
            hostItemValueLength = len(numList)
        except Exception:
            hostItemValueLength = 0
        try:
            hostItemValueSum = sum(numList)
        except Exception:
            hostItemValueSum = 0
        try:
            sla = (hostItemValueSum/hostItemValueLength) * 100
        except Exception:
            sla = "uncomputed"

# OS check
        itemOS = zabbixApi2.item.get(filter=itemFilterOS, host=x['host'], output='extend', selectHosts=['host','name'])
        try:
            for itemOsValue in itemOS:
                osVal = str(itemOsValue['lastvalue'])
                osVal = osVal.split(" ")
                osVal1 = str(osVal[0])
                osVal2 = str(osVal[1])
                osVal3 = str(osVal[2])
                osValAll = str(osVal1 + " " + osVal2 + " " + osVal3)
        except Exception:
            osValAll = "Undefined"

# Progress bar
        bar.next()

# Add all info into report
        ws2.append([hostid] + [host] + [name] + [status] + [ip] + [port] + [type] + [osValAll] + [avail] + [sla] + [inventory] + [groups] + [templates] + [itemsStr] + [triggersStr])

    bar.finish()

    print("|----- Collect info and add into report - done -----|")

# Let's configurate the report
# Merge cells for title
    maxCol = ws2.max_column
    minRow = ws2.min_row
    maxRow = ws2.max_row
    maxColLetter = get_column_letter(maxCol)
    refCell = maxColLetter+str(maxRow)

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=maxCol)

### Styles ###
# Column width
    for col in ws2.iter_cols(min_row=2, max_row=hostCount+2, min_col=1, max_col=10):
         max_length = 10
         column = col[0].column_letter
         for cell in col:
             try: 
                 if len(str(cell.value)) > max_length:
                     #max_length = len(str(cell.value))
                     max_length = 15
             except:
                 continue
         adjusted_width = (max_length)
         ws2.column_dimensions[column].width = adjusted_width

    for col in ws2.iter_cols(min_row=2, max_row=hostCount+2, min_col=11, max_col=13):
         max_length = 15
         column = col[0].column_letter
         for cell in col:
             try: 
                 if len(str(cell.value)) > max_length:
                     #max_length = len(str(cell.value))
                     max_length = 30
             except:
                 continue
         adjusted_width = (max_length)
         ws2.column_dimensions[column].width = adjusted_width


    for col in ws2.iter_cols(min_row=2, max_row=hostCount+2, min_col=14, max_col=15):
         max_length = 50
         column = col[0].column_letter
         for cell in col:
             try: 
                 if len(str(cell.value)) > max_length:
                     #max_length = len(str(cell.value))
                     max_length = 100
             except:
                 continue
         adjusted_width = (max_length)
         ws2.column_dimensions[column].width = adjusted_width


# Alignment style
    alignment=Alignment(
                    horizontal='left',
                    vertical='top',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=False,
                    indent=0
                       )

# Font styles
    fontTitle = Font(
            name='Bahnschrift SemiBold',
            size=14,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000'
                )

    fontHeaders = Font(
            name='Bahnschrift',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000'
                )

    fontCells = Font(
            name='Bahnschrift Light',
            size=9,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000'
                )

# Fill cells style
    fillTitle = PatternFill(fill_type='solid', fgColor='FFCBA4')
    fillHeaders = PatternFill(fill_type='solid', fgColor='96c8a2')
    fillCells = PatternFill(fill_type='solid', fgColor='addfad')
    fillAvail = PatternFill(fill_type='solid', fgColor='32CD32')
    fillUnreach = PatternFill(fill_type='solid', fgColor='CD5C5C')

# Borders style
    borderTitle = Side(border_style="thick", color="2f4f4f")
    borderHeader = Side(border_style="medium", color="2f4f4f")
    borderCells = Side(border_style="thin", color="2f4f4f")

# Title cell
    ws2['A1'].font = fontTitle
    ws2['A1'].fill = fillTitle
    ws2['A1'].alignment = alignment

# Headers
    for cells in ws2.iter_cols(min_row=2, max_row=2):
        for cell in cells:
            cell.font = fontHeaders
            cell.fill = fillHeaders
            cell.alignment = alignment
            cell.border = Border(top=borderHeader, bottom=borderHeader, left=borderHeader, right=borderHeader) 

# Fill cells skipping one
    for rows in range(minRow + 2, maxRow + 2, 2):
        for cells in ws2.iter_cols(min_row=rows, max_row=rows):
            for cell in cells:
                cell.fill = fillCells

# Data cells
    for cells in ws2.iter_cols(min_row=3, max_row=hostCount+2):
        for cell in cells:
            styleObj = cell.coordinate
            ws2[f'{styleObj}'].alignment = alignment
            ws2[f'{styleObj}'].border = Border(top=borderCells, bottom=borderCells, left=borderCells, right=borderCells)
            if cell.value == "available":
                cell.fill = fillAvail
            elif cell.value == "unreachable":
                cell.fill = fillUnreach

    for i in range(1, ws2.max_row+1):
        ws2.row_dimensions[i].height = 30

    ws2.page_setup.scale = 80

    print("|----- Configure report cells style - done -----|")


# Log out from Zabbix API
    zabbixApi2.user.logout()

else:
    print("Incorrect input. Try again.")

# Save file
wb.save(f"Zabbix_compare.xlsx")
    
print("|----- Save file - done -----|")
print("Execution time: ")

if platform.system().lower()!='windows':
    print("--- %s seconds ---" % round(time.time() - start_time, 2))
else:
    print("--- %s seconds ---" % round(time.clock() - start_time, 2))
